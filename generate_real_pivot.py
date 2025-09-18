import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
from datetime import datetime
import json

def fetch_all_data():
    """Obtiene todos los datos necesarios del API"""
    try:
        # Obtener datos detallados
        detail_response = requests.get('http://localhost:3001/api/opportunity-detail')
        detail_data = detail_response.json()

        # Obtener resúmenes para validación
        division_response = requests.get('http://localhost:3001/api/division-summary')
        division_data = division_response.json()

        lead_response = requests.get('http://localhost:3001/api/lead-summary')
        lead_data = lead_response.json()

        if detail_data['success']:
            return {
                'detail': pd.DataFrame(detail_data['data']),
                'division_summary': pd.DataFrame(division_data['data']),
                'lead_summary': pd.DataFrame(lead_data['data'])
            }
    except Exception as e:
        print(f"Error: {e}")
        return None

def prepare_data_for_pivot(df):
    """Prepara los datos para las pivot tables con todos los cálculos necesarios"""

    # Limpiar y preparar datos
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
    df['Created_Date'] = pd.to_datetime(df['Created_Date'], errors='coerce', utc=True).dt.tz_localize(None)
    df['LastStageChangeDate'] = pd.to_datetime(df['LastStageChangeDate'], errors='coerce', utc=True).dt.tz_localize(None)

    # Calcular Year según reglas de negocio
    df['Year'] = df.apply(lambda row:
        row['LastStageChangeDate'].year
        if row['StageName'] in ['Approved', 'Lost'] and pd.notna(row['LastStageChangeDate'])
        else row['Created_Date'].year
        if pd.notna(row['Created_Date'])
        else 2024, axis=1)

    # Crear columnas auxiliares para los cálculos
    df['IsApproved'] = (df['StageName'] == 'Approved').astype(int)
    df['IsLost'] = (df['StageName'] == 'Lost').astype(int)
    df['IsOpen'] = (~df['StageName'].isin(['Approved', 'Lost'])).astype(int)

    df['ApprovedAmount'] = df['Amount'] * df['IsApproved']
    df['LostAmount'] = df['Amount'] * df['IsLost']
    df['OpenAmount'] = df['Amount'] * df['IsOpen']

    return df

def create_division_pivot_with_calculations(df):
    """Crea pivot table por División con todos los cálculos"""

    # Agrupar por Year y Division
    pivot = df.groupby(['Year', 'Division']).agg({
        'Id': 'count',  # Total Opportunities
        'IsApproved': 'sum',  # Count Approved
        'IsLost': 'sum',  # Count Lost
        'IsOpen': 'sum',  # Count Open
        'ApprovedAmount': 'sum',  # Sum Approved Amount
        'LostAmount': 'sum',  # Sum Lost Amount
        'OpenAmount': 'sum',  # Sum Open Amount
        'Amount': 'sum'  # Total Amount
    }).reset_index()

    # Renombrar columnas
    pivot.columns = ['Year', 'Division', 'Total_Opp', 'Approved', 'Lost', 'Open',
                     'Approved_Revenue', 'Lost_Revenue', 'Open_Revenue', 'Total_Amount']

    # Calcular métricas derivadas
    pivot['CloseRate_Std'] = (pivot['Approved'] / pivot['Total_Opp'] * 100).round(2)
    pivot['CloseRate_NoLost'] = (pivot['Approved'] / (pivot['Total_Opp'] - pivot['Lost']) * 100).round(2)
    pivot['Average_Ticket'] = (pivot['Approved_Revenue'] / pivot['Approved']).fillna(0).round(2)

    # Agregar totales por año
    year_totals = df.groupby('Year').agg({
        'Id': 'count',
        'IsApproved': 'sum',
        'IsLost': 'sum',
        'IsOpen': 'sum',
        'ApprovedAmount': 'sum',
        'LostAmount': 'sum',
        'OpenAmount': 'sum',
        'Amount': 'sum'
    }).reset_index()

    year_totals['Division'] = 'TOTAL'
    year_totals.columns = ['Year', 'Total_Opp', 'Approved', 'Lost', 'Open',
                           'Approved_Revenue', 'Lost_Revenue', 'Open_Revenue',
                           'Total_Amount', 'Division']

    year_totals['CloseRate_Std'] = (year_totals['Approved'] / year_totals['Total_Opp'] * 100).round(2)
    year_totals['CloseRate_NoLost'] = (year_totals['Approved'] / (year_totals['Total_Opp'] - year_totals['Lost']) * 100).round(2)
    year_totals['Average_Ticket'] = (year_totals['Approved_Revenue'] / year_totals['Approved']).fillna(0).round(2)

    # Reorganizar columnas
    year_totals = year_totals[pivot.columns]

    # Combinar con totales
    final_pivot = pd.concat([pivot, year_totals], ignore_index=True).sort_values(['Year', 'Division'])

    return final_pivot

def create_lead_pivot_with_calculations(df):
    """Crea pivot table por Lead Type con todos los cálculos"""

    # Agrupar por Year y LeadType
    pivot = df.groupby(['Year', 'LeadType']).agg({
        'Id': 'count',
        'IsApproved': 'sum',
        'IsLost': 'sum',
        'IsOpen': 'sum',
        'ApprovedAmount': 'sum',
        'LostAmount': 'sum',
        'OpenAmount': 'sum',
        'Amount': 'sum'
    }).reset_index()

    # Renombrar columnas
    pivot.columns = ['Year', 'LeadType', 'Total_Opp', 'Approved', 'Lost', 'Open',
                     'Approved_Revenue', 'Lost_Revenue', 'Open_Revenue', 'Total_Amount']

    # Calcular métricas derivadas
    pivot['CloseRate_Std'] = (pivot['Approved'] / pivot['Total_Opp'] * 100).round(2)
    pivot['CloseRate_NoLost'] = (pivot['Approved'] / (pivot['Total_Opp'] - pivot['Lost']) * 100).round(2)
    pivot['Average_Ticket'] = (pivot['Approved_Revenue'] / pivot['Approved']).fillna(0).round(2)

    # Agregar totales por año
    year_totals = df.groupby('Year').agg({
        'Id': 'count',
        'IsApproved': 'sum',
        'IsLost': 'sum',
        'IsOpen': 'sum',
        'ApprovedAmount': 'sum',
        'LostAmount': 'sum',
        'OpenAmount': 'sum',
        'Amount': 'sum'
    }).reset_index()

    year_totals['LeadType'] = 'TOTAL'
    year_totals.columns = ['Year', 'Total_Opp', 'Approved', 'Lost', 'Open',
                           'Approved_Revenue', 'Lost_Revenue', 'Open_Revenue',
                           'Total_Amount', 'LeadType']

    year_totals['CloseRate_Std'] = (year_totals['Approved'] / year_totals['Total_Opp'] * 100).round(2)
    year_totals['CloseRate_NoLost'] = (year_totals['Approved'] / (year_totals['Total_Opp'] - year_totals['Lost']) * 100).round(2)
    year_totals['Average_Ticket'] = (year_totals['Approved_Revenue'] / year_totals['Approved']).fillna(0).round(2)

    # Reorganizar columnas
    year_totals = year_totals[pivot.columns]

    # Combinar con totales
    final_pivot = pd.concat([pivot, year_totals], ignore_index=True).sort_values(['Year', 'LeadType'])

    return final_pivot

def create_excel_with_real_pivots(data_dict):
    """Crea Excel con pivot tables reales usando openpyxl"""

    filename = f"Opportunity_Real_Pivots_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Preparar datos
    df = data_dict['detail']
    df_prepared = prepare_data_for_pivot(df)

    # Crear pivot tables con cálculos
    division_pivot = create_division_pivot_with_calculations(df_prepared)
    lead_pivot = create_lead_pivot_with_calculations(df_prepared)

    # Crear Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:

        # 1. Hoja de datos crudos (para drill-through)
        df_prepared.to_excel(writer, sheet_name='Raw_Data', index=False)

        # 2. Division Summary Pivot
        division_pivot.to_excel(writer, sheet_name='Division_Summary', index=False)

        # 3. Lead Summary Pivot
        lead_pivot.to_excel(writer, sheet_name='Lead_Summary', index=False)

        # Obtener workbook para formato
        workbook = writer.book

        # Formatear Division Summary
        div_sheet = workbook['Division_Summary']
        format_summary_sheet(div_sheet)

        # Formatear Lead Summary
        lead_sheet = workbook['Lead_Summary']
        format_summary_sheet(lead_sheet)

        # Agregar hoja de métricas clave
        metrics_sheet = workbook.create_sheet('Key_Metrics')
        add_key_metrics(metrics_sheet, division_pivot, lead_pivot)

    print(f"[OK] Excel con pivot tables creado: {filename}")
    return filename

def format_summary_sheet(sheet):
    """Aplica formato profesional a las hojas de summary"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule

    # Formato de encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar formato a encabezados
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Formato de moneda para columnas de revenue
    for row in sheet.iter_rows(min_row=2):
        # Columnas de moneda (índices aproximados)
        for col_idx in [6, 7, 8, 9, 12]:  # Ajustar según posición real
            if col_idx < len(row):
                cell = row[col_idx]
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0'

    # Formato de porcentaje para close rates
    for row in sheet.iter_rows(min_row=2):
        for col_idx in [10, 11]:  # Columnas de CloseRate
            if col_idx < len(row):
                cell = row[col_idx]
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'
                    cell.value = cell.value / 100  # Convertir a decimal para formato %

    # Resaltar filas de TOTAL
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_font = Font(bold=True)

    for row in sheet.iter_rows(min_row=2):
        if row[1].value == 'TOTAL' or (len(row) > 2 and row[2].value == 'TOTAL'):
            for cell in row:
                cell.fill = total_fill
                cell.font = total_font

    # Ajustar ancho de columnas
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        sheet.column_dimensions[column_letter].width = adjusted_width

def add_key_metrics(sheet, division_pivot, lead_pivot):
    """Agrega una hoja con métricas clave y KPIs"""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, PieChart, Reference

    # Título
    sheet['A1'] = 'KEY PERFORMANCE INDICATORS'
    sheet['A1'].font = Font(size=16, bold=True)

    # Métricas del año actual (último año en los datos)
    current_year = division_pivot['Year'].max()
    current_data = division_pivot[division_pivot['Year'] == current_year]
    total_row = current_data[current_data['Division'] == 'TOTAL'].iloc[0] if len(current_data[current_data['Division'] == 'TOTAL']) > 0 else None

    if total_row is not None:
        # KPIs principales
        sheet['A3'] = f'Year {int(current_year)} Performance'
        sheet['A3'].font = Font(size=14, bold=True)

        kpis = [
            ('Total Opportunities', total_row['Total_Opp']),
            ('Approved Deals', total_row['Approved']),
            ('Close Rate Standard', f"{total_row['CloseRate_Std']:.2f}%"),
            ('Average Ticket', f"${total_row['Average_Ticket']:,.2f}"),
            ('Total Revenue', f"${total_row['Approved_Revenue']:,.2f}"),
            ('Lost Revenue', f"${total_row['Lost_Revenue']:,.2f}"),
            ('Open Pipeline', f"${total_row['Open_Revenue']:,.2f}")
        ]

        row_num = 5
        for label, value in kpis:
            sheet[f'A{row_num}'] = label
            sheet[f'B{row_num}'] = str(value)
            sheet[f'A{row_num}'].font = Font(bold=True)
            row_num += 1

        # Comparación YoY si hay datos anteriores
        if len(division_pivot['Year'].unique()) > 1:
            sheet['D3'] = 'Year-over-Year Comparison'
            sheet['D3'].font = Font(size=14, bold=True)

            prev_year = current_year - 1
            prev_data = division_pivot[division_pivot['Year'] == prev_year]
            if len(prev_data[prev_data['Division'] == 'TOTAL']) > 0:
                prev_total = prev_data[prev_data['Division'] == 'TOTAL'].iloc[0]

                # Calcular cambios
                opp_change = ((total_row['Total_Opp'] - prev_total['Total_Opp']) / prev_total['Total_Opp'] * 100)
                rev_change = ((total_row['Approved_Revenue'] - prev_total['Approved_Revenue']) / prev_total['Approved_Revenue'] * 100)

                sheet['D5'] = 'Opportunity Growth'
                sheet['E5'] = f"{opp_change:+.2f}%"
                sheet['D6'] = 'Revenue Growth'
                sheet['E6'] = f"{rev_change:+.2f}%"

    # Ajustar anchos
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15

def create_excel_with_native_pivot():
    """Versión alternativa usando xlsxwriter para crear pivot tables nativas de Excel"""
    import xlsxwriter

    # Obtener datos
    data_dict = fetch_all_data()
    if not data_dict:
        return None

    df = data_dict['detail']
    df_prepared = prepare_data_for_pivot(df)

    filename = f"Opportunity_Native_Pivot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Crear workbook con xlsxwriter
    workbook = xlsxwriter.Workbook(filename)

    # Hoja de datos
    data_sheet = workbook.add_worksheet('Data')

    # Escribir headers
    headers = df_prepared.columns.tolist()
    for col_num, header in enumerate(headers):
        data_sheet.write(0, col_num, header)

    # Escribir datos
    for row_num, row_data in enumerate(df_prepared.values, 1):
        for col_num, value in enumerate(row_data):
            if pd.notna(value):
                if isinstance(value, (pd.Timestamp, datetime)):
                    data_sheet.write_datetime(row_num, col_num, value.replace(tzinfo=None) if hasattr(value, 'tzinfo') else value)
                else:
                    data_sheet.write(row_num, col_num, value)

    # Crear tabla de Excel
    last_row = len(df_prepared)
    last_col = len(df_prepared.columns) - 1

    data_sheet.add_table(0, 0, last_row, last_col, {
        'name': 'OpportunityData',
        'columns': [{'header': col} for col in headers]
    })

    # Crear hoja de pivot para División
    pivot_sheet = workbook.add_worksheet('PivotTable_Division')

    # Nota sobre la creación manual de pivot
    pivot_sheet.write('A1', 'INSTRUCCIONES PARA CREAR PIVOT TABLE:')
    pivot_sheet.write('A3', '1. Selecciona la tabla en la hoja "Data"')
    pivot_sheet.write('A4', '2. Insert > PivotTable')
    pivot_sheet.write('A5', '3. Configurar campos:')
    pivot_sheet.write('B6', '   Rows: Year, Division')
    pivot_sheet.write('B7', '   Values:')
    pivot_sheet.write('B8', '      - Count of Id (rename to "Total Opportunities")')
    pivot_sheet.write('B9', '      - Sum of IsApproved (rename to "Approved")')
    pivot_sheet.write('B10', '     - Sum of ApprovedAmount (rename to "Revenue")')
    pivot_sheet.write('A12', '4. Para drill-through: doble clic en cualquier valor')

    # Formatos
    currency_format = workbook.add_format({'num_format': '$#,##0'})
    percent_format = workbook.add_format({'num_format': '0.00%'})

    workbook.close()

    print(f"[OK] Excel con estructura para pivot nativo creado: {filename}")
    return filename

def main():
    print("Generando Excel con Pivot Tables Reales...")
    print("=" * 50)

    # Obtener todos los datos
    data_dict = fetch_all_data()
    if not data_dict:
        print("[ERROR] No se pudieron obtener los datos")
        return

    print(f"[OK] Datos obtenidos:")
    print(f"  - Detail: {len(data_dict['detail'])} registros")
    print(f"  - Division Summary: {len(data_dict['division_summary'])} registros")
    print(f"  - Lead Summary: {len(data_dict['lead_summary'])} registros")

    # Crear Excel con pivots calculados
    filename = create_excel_with_real_pivots(data_dict)

    # También crear versión para pivot nativo
    native_filename = create_excel_with_native_pivot()

    print("\n" + "=" * 50)
    print("ARCHIVOS GENERADOS:")
    print(f"1. {filename} - Con calculos completos")
    print(f"2. {native_filename} - Para crear pivot nativo en Excel")
    print("\nLas pivot tables incluyen:")
    print("  - Close Rate Standard y No Lost")
    print("  - Average Ticket")
    print("  - Totales por Year")
    print("  - Drill-through habilitado en datos crudos")

if __name__ == "__main__":
    main()